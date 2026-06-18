#!/usr/bin/env python3
"""
Oracle Cloud Object Storage — TCP Latency Probe

Measures TCP handshake latency (port 443) to all 42 Oracle Cloud
Object Storage endpoints worldwide. Outputs:
  - oci_latency_report.xlsx   (9-sheet Excel workbook)
  - oci_latency_chart.html    (interactive ECharts visualization)

Usage:
    python oci_latency_probe.py
"""

import json
import socket
import statistics
import sys
import time
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Dependencies
# ---------------------------------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    print("Installing openpyxl ...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

# ===========================================================================
#  CONFIG
# ===========================================================================

PROBE_COUNT = 30
PORT = 443
TIMEOUT = 5.0

# (region_code, name_cn, name_en, host)
ENDPOINTS = [
    # -- Asia-Pacific --
    ("ap-osaka-1",       "日本中部大阪",           "Osaka, Japan",                  "objectstorage.ap-osaka-1.oraclecloud.com"),
    ("ap-tokyo-1",       "日本东部东京",           "Tokyo, Japan",                  "objectstorage.ap-tokyo-1.oraclecloud.com"),
    ("ap-seoul-1",       "韩国中部首尔",           "Seoul, South Korea",            "objectstorage.ap-seoul-1.oraclecloud.com"),
    ("ap-chuncheon-1",   "韩国北部春川",           "Chuncheon, South Korea",        "objectstorage.ap-chuncheon-1.oraclecloud.com"),
    ("ap-singapore-1",   "新加坡",                 "Singapore",                     "objectstorage.ap-singapore-1.oraclecloud.com"),
    ("ap-singapore-2",   "新加坡西",               "Singapore West",                "objectstorage.ap-singapore-2.oraclecloud.com"),
    ("ap-mumbai-1",      "印度西部孟买",           "Mumbai, India",                 "objectstorage.ap-mumbai-1.oraclecloud.com"),
    ("ap-hyderabad-1",   "印度南部海得拉巴",       "Hyderabad, India",              "objectstorage.ap-hyderabad-1.oraclecloud.com"),
    ("ap-batam-1",       "印度尼西亚巴淡",         "Batam, Indonesia",              "objectstorage.ap-batam-1.oraclecloud.com"),
    ("ap-sydney-1",      "澳大利亚东部悉尼",       "Sydney, Australia",             "objectstorage.ap-sydney-1.oraclecloud.com"),
    ("ap-melbourne-1",   "澳大利亚东南部墨尔本",   "Melbourne, Australia",          "objectstorage.ap-melbourne-1.oraclecloud.com"),
    # -- North America --
    ("us-ashburn-1",     "美国东部阿什本",         "Ashburn, USA",                  "objectstorage.us-ashburn-1.oraclecloud.com"),
    ("us-phoenix-1",     "美国西部凤凰城",         "Phoenix, USA",                  "objectstorage.us-phoenix-1.oraclecloud.com"),
    ("us-sanjose-1",     "美国西部圣何塞",         "San Jose, USA",                 "objectstorage.us-sanjose-1.oraclecloud.com"),
    ("us-chicago-1",     "美国中西部芝加哥",       "Chicago, USA",                  "objectstorage.us-chicago-1.oraclecloud.com"),
    ("ca-toronto-1",     "加拿大东南部多伦多",     "Toronto, Canada",               "objectstorage.ca-toronto-1.oraclecloud.com"),
    ("ca-montreal-1",    "加拿大东南部蒙特利尔",   "Montreal, Canada",              "objectstorage.ca-montreal-1.oraclecloud.com"),
    ("mx-monterrey-1",   "墨西哥东北部蒙特雷",     "Monterrey, Mexico",             "objectstorage.mx-monterrey-1.oraclecloud.com"),
    ("mx-queretaro-1",   "墨西哥中部克雷塔罗",     "Querétaro, Mexico",             "objectstorage.mx-queretaro-1.oraclecloud.com"),
    # -- South America --
    ("sa-saopaulo-1",    "巴西东部圣保罗",         "São Paulo, Brazil",             "objectstorage.sa-saopaulo-1.oraclecloud.com"),
    ("sa-vinhedo-1",     "巴西南部维涅杜",         "Vinhedo, Brazil",               "objectstorage.sa-vinhedo-1.oraclecloud.com"),
    ("sa-santiago-1",    "智利中部圣地亚哥",       "Santiago, Chile",               "objectstorage.sa-santiago-1.oraclecloud.com"),
    ("sa-valparaiso-1",  "智利西部瓦尔帕莱索",     "Valparaíso, Chile",             "objectstorage.sa-valparaiso-1.oraclecloud.com"),
    ("sa-bogota-1",      "哥伦比亚中部波哥大",     "Bogotá, Colombia",              "objectstorage.sa-bogota-1.oraclecloud.com"),
    # -- Europe --
    ("uk-london-1",      "英国南部伦敦",           "London, UK",                    "objectstorage.uk-london-1.oraclecloud.com"),
    ("uk-cardiff-1",     "英国西部加的夫",         "Cardiff, UK",                   "objectstorage.uk-cardiff-1.oraclecloud.com"),
    ("eu-paris-1",       "法国中部巴黎",           "Paris, France",                 "objectstorage.eu-paris-1.oraclecloud.com"),
    ("eu-marseille-1",   "法国南部马赛",           "Marseille, France",             "objectstorage.eu-marseille-1.oraclecloud.com"),
    ("eu-frankfurt-1",   "德国中部法兰克福",       "Frankfurt, Germany",            "objectstorage.eu-frankfurt-1.oraclecloud.com"),
    ("eu-zurich-1",      "瑞士北部苏黎世",         "Zurich, Switzerland",           "objectstorage.eu-zurich-1.oraclecloud.com"),
    ("eu-turin-1",       "意大利北部都灵",         "Turin, Italy",                  "objectstorage.eu-turin-1.oraclecloud.com"),
    ("eu-milan-1",       "意大利西北部米兰",       "Milan, Italy",                  "objectstorage.eu-milan-1.oraclecloud.com"),
    ("eu-madrid-1",      "西班牙中部马德里",       "Madrid, Spain",                 "objectstorage.eu-madrid-1.oraclecloud.com"),
    ("eu-madrid-3",      "西班牙中部马德里3",      "Madrid 3, Spain",               "objectstorage.eu-madrid-3.oraclecloud.com"),
    ("il-jerusalem-1",   "以色列中部耶路撒冷",     "Jerusalem, Israel",             "objectstorage.il-jerusalem-1.oraclecloud.com"),
    ("eu-stockholm-1",   "瑞典中部斯德哥尔摩",     "Stockholm, Sweden",             "objectstorage.eu-stockholm-1.oraclecloud.com"),
    ("eu-amsterdam-1",   "荷兰西北部阿姆斯特丹",   "Amsterdam, Netherlands",        "objectstorage.eu-amsterdam-1.oraclecloud.com"),
    # -- Middle East --
    ("me-dubai-1",       "阿联酋迪拜",             "Dubai, UAE",                    "objectstorage.me-dubai-1.oraclecloud.com"),
    ("me-abudhabi-1",    "阿联酋阿布扎比",         "Abu Dhabi, UAE",                "objectstorage.me-abudhabi-1.oraclecloud.com"),
    ("me-jeddah-1",      "沙特阿拉伯西部吉达",     "Jeddah, Saudi Arabia",          "objectstorage.me-jeddah-1.oraclecloud.com"),
    ("me-riyadh-1",      "沙特阿拉伯中部利雅得",   "Riyadh, Saudi Arabia",          "objectstorage.me-riyadh-1.oraclecloud.com"),
    # -- Africa --
    ("af-johannesburg-1","南非中部约翰内斯堡",     "Johannesburg, South Africa",    "objectstorage.af-johannesburg-1.oraclecloud.com"),
]

REGION_GROUPS = [
    ("亚太地区 (APAC)",      ("ap",),           ["#5470C6","#73A0FA","#91CC75","#FAC858","#EE6666","#3BA272","#FC8452","#9A60B4","#EA7CCC","#48B8D0","#6E7074"]),
    ("北美地区 (NA)",        ("us","ca","mx"),   ["#C1232B","#D7504B","#E87C7E","#F19C9C","#D48265","#E69C87","#F4B7A4","#B5CAA0"]),
    ("南美地区 (SA)",        ("sa",),            ["#749F83","#91B493","#A8C8A8","#C4DBC4","#94A87C"]),
    ("欧洲地区 (EU)",        ("uk","eu","il"),   ["#6F4E9B","#8B6BAE","#A78BBF","#C4ABD2","#9B59B6","#A569BD","#B07CC6","#C39BD3","#D2B4DE","#A7C5EB","#B8D3F0","#C9E0F5","#DAECFA"]),
    ("中东地区 (ME)",        ("me",),            ["#F4A460","#F6B87A","#F8CC94","#FADFAD"]),
    ("非洲地区 (AF)",        ("af",),            ["#8B7765"]),
]

# ===========================================================================
#  NETWORK PROBE
# ===========================================================================

def tcp_latency(host, port=PORT, timeout=TIMEOUT):
    """Measure TCP handshake latency (ms). Returns None on failure."""
    try:
        t0 = time.perf_counter()
        with socket.create_connection((host, port), timeout=timeout):
            return round((time.perf_counter() - t0) * 1000, 2)
    except Exception:
        return None

def probe_host(host, count=PROBE_COUNT):
    """Probe host `count` times, return latencies in ms."""
    latencies = []
    for _ in range(count):
        lat = tcp_latency(host)
        if lat is not None:
            latencies.append(lat)
        time.sleep(0.5)
    return latencies

def trimmed_mean(data, trim_pct=0.10):
    """Mean after trimming top & bottom `trim_pct` extremes."""
    if len(data) < 3:
        return None
    s = sorted(data)
    k = max(1, int(len(s) * trim_pct))
    trimmed = s[k: len(s) - k]
    return round(statistics.mean(trimmed), 2) if trimmed else None

# ===========================================================================
#  EXCEL BUILDER
# ===========================================================================

def _styles():
    return dict(
        hfont=Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF"),
        hfill=PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        halign=Alignment(horizontal="center", vertical="center", wrap_text=True),
        calign=Alignment(horizontal="center", vertical="center"),
        lalign=Alignment(horizontal="left", vertical="center"),
        border=Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin")),
        green=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        yellow=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        red=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        tfont=Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79"),
    )

def _latency_fill(val, st):
    """Return green/yellow/red fill for a latency value."""
    if val is None:
        return None
    return st["green"] if val < 100 else (st["yellow"] if val < 300 else st["red"])

def _set_fill(cell, fill):
    """Set cell fill only if fill is not None (avoids openpyxl crash on None)."""
    if fill is not None:
        cell.fill = fill

def _loss_fill(lost, st):
    """Return green/yellow/red fill for lost count."""
    if lost == 0:
        return st["green"]
    elif lost < 15:
        return st["yellow"]
    return st["red"]

def _build_sheet(ws, title, results, st, merge_end="L"):
    """Write a summary sheet (main or per-region)."""
    ws.merge_cells(f"A1:{merge_end}1")
    ws["A1"].value = title
    ws["A1"].font = st["tfont"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [
        "Region Code", "Region Name (CN)", "Region Name (EN)", "Hostname",
        "Sent", "Received", "Lost", "Loss %",
        "Min (ms)", "Max (ms)", "Avg (ms)", "Trimmed Avg (ms)",
    ]
    widths = [18, 24, 28, 48, 8, 10, 8, 10, 12, 12, 12, 18]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = st["hfont"]; cell.fill = st["hfill"]
        cell.alignment = st["halign"]; cell.border = st["border"]
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 28

    for r, rec in enumerate(results, 3):
        vals = [
            rec["region"], rec["name_cn"], rec["name_en"], rec["host"],
            rec["sent"], rec["received"], rec["lost"], rec["loss_pct"],
            rec["min_ms"], rec["max_ms"], rec["avg_ms"], rec["trimmed_avg_ms"],
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val if val is not None else "N/A")
            cell.border = st["border"]
            cell.alignment = st["lalign"] if c <= 4 else st["calign"]

        # Color-code: Lost (col 7), Loss% (col 8), Min (col 9), Max (col 10), Avg (col 11), Trimmed (col 12)
        ws.cell(row=r, column=7).fill = _loss_fill(rec["lost"], st)
        ws.cell(row=r, column=8).fill = _loss_fill(rec["lost"], st)
        _set_fill(ws.cell(row=r, column=9), _latency_fill(rec["min_ms"], st))
        _set_fill(ws.cell(row=r, column=10), _latency_fill(rec["max_ms"], st))
        _set_fill(ws.cell(row=r, column=11), _latency_fill(rec["avg_ms"], st))
        _set_fill(ws.cell(row=r, column=12), _latency_fill(rec["trimmed_avg_ms"], st))

    ws.freeze_panes = "A3"
    if results:
        ws.auto_filter.ref = f"A2:{merge_end}{2 + len(results)}"

def _build_raw_sheet(wb, results, st):
    ws = wb.create_sheet("Raw Data")
    n = PROBE_COUNT
    last_col = get_column_letter(4 + n)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].value = f"Raw Probe Data — {n} TCP connections per endpoint (ms)"
    ws["A1"].font = st["tfont"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    all_h = ["Region Code", "Region Name (CN)", "Region Name (EN)", "Hostname"] + [f"P{i}" for i in range(1, n + 1)]
    for c, h in enumerate(all_h, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = st["hfont"]; cell.fill = st["hfill"]
        cell.alignment = st["halign"]; cell.border = st["border"]
    ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28; ws.column_dimensions["D"].width = 48
    for c in range(5, 5 + n):
        ws.column_dimensions[get_column_letter(c)].width = 9
    ws.row_dimensions[2].height = 28

    for r, rec in enumerate(results, 3):
        for c, val in enumerate([rec["region"], rec["name_cn"], rec["name_en"], rec["host"]], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = st["border"]; cell.alignment = st["lalign"]
        lats = rec.get("latencies", [])
        for i in range(n):
            val = lats[i] if i < len(lats) else None
            cell = ws.cell(row=r, column=5 + i, value=val if val is not None else "N/A")
            cell.border = st["border"]; cell.alignment = st["calign"]
            if val is not None:
                _set_fill(cell, _latency_fill(val, st))

    ws.freeze_panes = "E3"
    if results:
        ws.auto_filter.ref = f"A2:{last_col}{2 + len(results)}"

def _build_summary_sheet(wb, results, st):
    ws = wb.create_sheet("Summary by Region")
    region_map = {
        "ap":"Asia-Pacific","us":"North America","ca":"North America","mx":"North America",
        "sa":"South America","uk":"Europe","eu":"Europe","il":"Europe",
        "me":"Middle East","af":"Africa",
    }
    order = ["Asia-Pacific","North America","South America","Europe","Middle East","Africa"]

    summary = {r:[] for r in order}
    for rec in results:
        grp = region_map.get(rec["region"].split("-")[0], "Other")
        if rec["avg_ms"] is not None:
            summary[grp].append(rec["avg_ms"])

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Summary by Geographic Region"
    ws["A1"].font = st["tfont"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for c,(h,w) in enumerate(zip(["Region","Endpoints Tested","Best Avg (ms)","Worst Avg (ms)"],[24,20,18,18]),1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = st["hfont"]; cell.fill = st["hfill"]
        cell.alignment = st["halign"]; cell.border = st["border"]
        ws.column_dimensions[get_column_letter(c)].width = w

    for r, region in enumerate(order, 3):
        avgs = summary[region]
        for c, val in enumerate([region, len(avgs),
            round(min(avgs),2) if avgs else "N/A",
            round(max(avgs),2) if avgs else "N/A"], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = st["border"]; cell.alignment = st["calign"]
    ws.freeze_panes = "A3"

def build_excel(results, output_path):
    wb = Workbook(); st = _styles()

    ws = wb.active; ws.title = "OCI TCP Latency Report"
    _build_sheet(ws, f"Oracle Cloud Object Storage — TCP Latency Report (port {PORT}, {PROBE_COUNT} probes each)", results, st)

    for sheet_name, prefixes, _ in REGION_GROUPS:
        region_results = [r for r in results if r["region"].split("-")[0] in prefixes]
        ws_r = wb.create_sheet(sheet_name)
        _build_sheet(ws_r, f"{sheet_name} — Oracle Cloud Object Storage TCP Latency", region_results, st)

    _build_raw_sheet(wb, results, st)
    _build_summary_sheet(wb, results, st)
    wb.save(output_path)
    return output_path

# ===========================================================================
#  ECHARTS HTML CHART
# ===========================================================================

def build_chart_html(results, output_path):
    """Generate a self-contained HTML with an interactive ECharts line chart."""

    # Build chart data grouped by region
    chart_data = []
    for sheet_name, prefixes, colors in REGION_GROUPS:
        eps = []
        region_results = [r for r in results if r["region"].split("-")[0] in prefixes]
        for i, rec in enumerate(region_results):
            eps.append({
                "id": rec["region"],
                "label": rec["name_en"],
                "label_cn": rec["name_cn"],
                "color": colors[i % len(colors)],
                "latencies": rec.get("latencies", []),
                "avg": rec.get("avg_ms"),
            })
        chart_data.append({"name": sheet_name, "endpoints": eps})

    data_json = json.dumps(chart_data, ensure_ascii=False)

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>OCI Latency Probe — ECharts</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.1/dist/echarts.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:"Segoe UI","Microsoft YaHei",sans-serif;background:#1a1a2e;color:#eee;overflow:hidden;height:100vh}}
.header{{background:linear-gradient(135deg,#0f0f23 0%,#1a1a3e 100%);padding:14px 24px;border-bottom:1px solid #2a2a4a;display:flex;align-items:center;justify-content:space-between}}
.header h1{{font-size:18px;font-weight:600;color:#e0e0ff}}
.header .sub{{font-size:12px;color:#888;margin-top:2px}}
.header .stats{{display:flex;gap:20px;font-size:12px;color:#aaa}}
.header .stats span{{color:#7eb8ff}}
.main{{display:flex;height:calc(100vh - 62px)}}
.sidebar{{width:300px;min-width:300px;background:#16162b;overflow-y:auto;border-right:1px solid #2a2a4a;padding:8px 0}}
.sidebar h3{{font-size:11px;color:#666;text-transform:uppercase;letter-spacing:2px;padding:10px 16px 6px}}
.region-group{{border-bottom:1px solid #1f1f3a}}
.region-hdr{{display:flex;align-items:center;padding:9px 14px;cursor:pointer;user-select:none;font-size:13px;font-weight:600;color:#ccc;transition:background .15s}}
.region-hdr:hover{{background:#1e1e38}}
.region-hdr .arrow{{font-size:10px;margin-right:8px;transition:transform .2s;width:12px;color:#666}}
.region-hdr .arrow.open{{transform:rotate(90deg)}}
.region-hdr input{{margin-right:8px;accent-color:#5470C6;transform:scale(1.05)}}
.region-hdr .badge{{font-size:10px;background:#2a2a4a;color:#888;border-radius:10px;padding:1px 8px;margin-left:auto}}
.ep-list{{display:none;padding:0 0 6px}}
.ep-list.open{{display:block}}
.ep-row{{display:flex;align-items:center;padding:4px 14px 4px 44px;font-size:12px;color:#aaa;cursor:pointer;transition:background .1s}}
.ep-row:hover{{background:#1e1e38}}
.ep-row .dot{{width:8px;height:8px;border-radius:50%;margin-right:8px;flex-shrink:0}}
.ep-row input{{margin-right:8px;accent-color:#5470C6;transform:scale(.9)}}
.ep-row .cn{{color:#666;margin-left:6px;font-size:11px}}
.chart-area{{flex:1;display:flex;flex-direction:column;min-width:0}}
.toolbar{{display:flex;gap:6px;padding:8px 16px;background:#16162b;border-bottom:1px solid #2a2a4a;flex-wrap:wrap}}
.toolbar button{{font-size:11px;padding:5px 12px;border:1px solid #3a3a5a;border-radius:4px;background:transparent;color:#aaa;cursor:pointer;transition:all .15s;white-space:nowrap}}
.toolbar button:hover{{background:#2a2a5a;color:#ddd;border-color:#5470C6}}
.toolbar button.on{{background:#5470C6;color:#fff;border-color:#5470C6}}
#chart{{flex:1;min-height:0;position:relative}}
#chart::after{{content:'';position:absolute;right:0;top:0;bottom:0;width:40px;background:linear-gradient(to right,transparent,#1a1a2e);pointer-events:none;z-index:10}}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>OCI Object Storage — TCP Latency Probe</h1>
    <div class="sub">42 endpoints · {PROBE_COUNT} probes each · port {PORT} · drag to zoom · scroll to navigate</div>
  </div>
  <div class="stats">
    <div>Fastest <span id="stat-best">--</span></div>
    <div>Slowest <span id="stat-worst">--</span></div>
    <div>Median <span id="stat-median">--</span></div>
  </div>
</div>
<div class="main">
  <div class="sidebar" id="sidebar">
    <h3>Toggle Regions &amp; Endpoints</h3>
    <div id="region-list"></div>
  </div>
  <div class="chart-area">
    <div class="toolbar">
      <button onclick="selAll()">Select All</button>
      <button onclick="selNone()">Deselect All</button>
      <button onclick="selBest(5)">Top 5 Fastest</button>
      <button onclick="selWorst(5)">Top 5 Slowest</button>
      <button onclick="byRegion()">Region Avg</button>
      <button id="btn-clip" onclick="toggleClip()">Clip Extremes</button>
      <button onclick="resetChart()">Reset</button>
    </div>
    <div id="chart"></div>
  </div>
</div>

<script>
const DATA = {data_json};
const PC = {PROBE_COUNT};

// ---- Build sidebar ----
const regionList = document.getElementById('region-list');
const seriesMeta = []; // {{id, regionIdx, epIdx, avg}}

DATA.forEach((region, ri) => {{
  const div = document.createElement('div');
  div.className = 'region-group';
  div.innerHTML = `
    <div class="region-hdr" onclick="toggleRegion(${{ri}})">
      <span class="arrow open" id="arr_${{ri}}">&#9654;</span>
      <input type="checkbox" checked onchange="toggleRegionSeries(${{ri}},this.checked);event.stopPropagation()">
      <span>${{region.name}}</span>
      <span class="badge">${{region.endpoints.length}}</span>
    </div>
    <div class="ep-list open" id="list_${{ri}}">
      ${{region.endpoints.map((ep, ei) => {{
        seriesMeta.push({{id:ep.id, regionIdx:ri, epIdx:ei, avg:ep.avg}});
        return `
          <div class="ep-row">
            <span class="dot" style="background:${{ep.color}}"></span>
            <input type="checkbox" checked onchange="toggleEp('${{ep.id}}');event.stopPropagation()">
            <span style="cursor:pointer" onclick="toggleEp('${{ep.id}}')">${{ep.label}}</span>
            <span class="cn">${{ep.label_cn}}</span>
          </div>`;
      }}).join('')}}
    </div>`;
  regionList.appendChild(div);
}});

// ---- Build ECharts series ----
const series = [];
const legendData = [];
DATA.forEach((region, ri) => {{
  region.endpoints.forEach((ep, ei) => {{
    const name = ep.label + ' (' + ep.label_cn + ')';
    legendData.push(name);
    series.push({{
      name: name,
      type: 'line',
      data: ep.latencies,
      smooth: true,
      symbol: 'none',
      lineStyle: {{ color: ep.color, width: 1.5 }},
      legendHoverLink: true,
    }});
  }});
  // Region average series (hidden by default)
  const regionAvgIds = [];
  const regionColors = ['#FF6B6B','#4ECDC4','#FFE66D','#A78BFA','#F97316','#10B981'];
  DATA.forEach((region, ri) => {{
    const rName = region.name;
    const eps = region.endpoints;
    const avgData = [];
    for (let p = 0; p < PC; p++) {{
      let sum = 0, cnt = 0;
      eps.forEach(ep => {{
        if (ep.latencies[p] != null) {{ sum += ep.latencies[p]; cnt++; }}
      }});
      avgData.push(cnt > 0 ? +(sum / cnt).toFixed(1) : null);
    }}
    const id = 'region_avg_' + ri;
    regionAvgIds.push(id);
    legendData.push('◆ ' + rName);
    series.push({{
      name: '◆ ' + rName,
      type: 'line',
      data: avgData,
      smooth: true,
      symbol: 'none',
      lineStyle: {{ color: regionColors[ri], width: 2.5, type: 'dashed' }},
      legendHoverLink: true,
    }});
  }});
}});

// ---- Init ECharts ----
const chartDom = document.getElementById('chart');
const chart = echarts.init(chartDom);

const option = {{
  backgroundColor: '#1a1a2e',
  tooltip: {{
    trigger: 'axis',
    backgroundColor: 'rgba(20,20,40,0.95)',
    borderColor: '#3a3a5a',
    textStyle: {{ color: '#ddd', fontSize: 12 }},
    formatter: function(params) {{
      let s = '<b>Probe ' + params[0].axisValue + '</b><br/>';
      params.forEach(p => {{
        if (p.value == null) return;
        s += '<span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:' + p.color + ';margin-right:6px"></span>';
        s += p.seriesName + ': <b>' + p.value.toFixed(1) + ' ms</b><br/>';
      }});
      return s;
    }}
  }},
  legend: {{
    type: 'scroll',
    orient: 'horizontal',
    left: 0,
    right: 0,
    bottom: 0,
    itemWidth: 12,
    itemHeight: 6,
    itemGap: 10,
    textStyle: {{ color: '#999', fontSize: 10 }},
    pageTextStyle: {{ color: '#777', fontSize: 10 }},
    pageIconColor: '#5470C6',
    pageIconInactiveColor: '#444',
    data: legendData,
    selected: legendData.reduce((acc, n) => {{ acc[n] = !n.startsWith('◆'); return acc; }}, {{}})
  }},
  grid: {{
    left: 55, right: 30, top: 50, bottom: 80
  }},
  xAxis: {{
    type: 'category',
    data: Array.from({{length: PC}}, (_,i) => i+1),
    name: 'Probe Number',
    nameTextStyle: {{ color: '#888' }},
    axisLine: {{ lineStyle: {{ color: '#333' }} }},
    axisTick: {{ show: false }},
    axisLabel: {{ color: '#777', fontSize: 11 }},
    splitLine: {{ show: false }}
  }},
  yAxis: {{
    type: 'value',
    name: 'Latency (ms)',
    nameTextStyle: {{ color: '#888' }},
    axisLine: {{ show: false }},
    axisTick: {{ show: false }},
    axisLabel: {{ color: '#777', fontSize: 11 }},
    splitLine: {{ lineStyle: {{ color: '#222', type: 'dashed' }} }}
  }},
  dataZoom: [
    {{
      type: 'slider',
      start: 0, end: 100,
      height: 24,
      bottom: 8,
      borderColor: '#333',
      backgroundColor: '#1a1a2e',
      dataBackground: {{
        lineStyle: {{ color: '#5470C6', opacity: 0.4 }},
        areaStyle: {{ color: '#5470C6', opacity: 0.1 }}
      }},
      selectedDataBackground: {{
        lineStyle: {{ color: '#5470C6' }},
        areaStyle: {{ color: '#5470C6', opacity: 0.2 }}
      }},
      handleStyle: {{ color: '#5470C6', borderColor: '#5470C6' }},
      textStyle: {{ color: '#888', fontSize: 10 }}
    }},
    {{
      type: 'inside',
      start: 0, end: 100
    }}
  ],
  toolbox: {{
    right: 30, top: 5,
    itemSize: 14,
    feature: {{
      saveAsImage: {{ title: 'Save', backgroundColor: '#1a1a2e' }},
      dataZoom: {{ title: {{ zoom: 'Zoom', back: 'Reset' }} }},
      restore: {{ title: 'Restore' }}
    }}
  }},
  series: series
}};

chart.setOption(option);

// ---- Stats ----
function updateStats() {{
  const visible = seriesMeta.filter(m => {{
    const name = DATA[m.regionIdx].endpoints[m.epIdx].label + ' (' + DATA[m.regionIdx].endpoints[m.epIdx].label_cn + ')';
    return chart.getOption().legend[0].selected[name] !== false;
  }});
  if (visible.length === 0) {{
    document.getElementById('stat-best').textContent = '--';
    document.getElementById('stat-worst').textContent = '--';
    document.getElementById('stat-median').textContent = '--';
    return;
  }}
  const avgs = visible.map(m => m.avg).filter(a => a != null);
  if (avgs.length === 0) return;
  avgs.sort((a,b)=>a-b);
  document.getElementById('stat-best').textContent = avgs[0].toFixed(1) + ' ms';
  document.getElementById('stat-worst').textContent = avgs[avgs.length-1].toFixed(1) + ' ms';
  document.getElementById('stat-median').textContent = avgs[Math.floor(avgs.length/2)].toFixed(1) + ' ms';
}}

// ---- Sidebar interactions ----
function toggleRegion(ri) {{
  const list = document.getElementById('list_' + ri);
  const arrow = document.getElementById('arr_' + ri);
  list.classList.toggle('open');
  arrow.classList.toggle('open');
}}

function toggleRegionSeries(ri, visible) {{
  const names = DATA[ri].endpoints.map(ep => ep.label + ' (' + ep.label_cn + ')');
  const sel = {{}};
  names.forEach(n => sel[n] = visible);
  chart.setOption({{ legend: {{ selected: sel }} }});
  // sync checkboxes
  const cbs = document.querySelectorAll('#list_' + ri + ' .ep-row input[type=checkbox]');
  cbs.forEach(cb => cb.checked = visible);
  updateStats();
  applyClip();
}}

function toggleEp(id) {{
  const meta = seriesMeta.find(m => m.id === id);
  if (!meta) return;
  const ep = DATA[meta.regionIdx].endpoints[meta.epIdx];
  const name = ep.label + ' (' + ep.label_cn + ')';
  const current = chart.getOption().legend[0].selected;
  const newVal = current[name] === false ? true : false;
  chart.setOption({{ legend: {{ selected: {{ [name]: newVal }} }} }});
  // sync checkbox
  const cb = document.querySelectorAll('#list_' + meta.regionIdx + ' .ep-row input[type=checkbox]')[meta.epIdx];
  if (cb) cb.checked = newVal;
  updateStats();
  applyClip();
}}

function selAll() {{
  const sel = {{}};
  legendData.forEach(n => sel[n] = !n.startsWith('◆'));
  chart.setOption({{ legend: {{ selected: sel }} }});
  document.querySelectorAll('input[type=checkbox]').forEach(cb => cb.checked = true);
  document.querySelector('button[onclick="byRegion()"]').classList.remove('on');
  updateStats();
  applyClip();
}}

function selNone() {{
  const sel = {{}};
  legendData.forEach(n => sel[n] = false);
  chart.setOption({{ legend: {{ selected: sel }} }});
  document.querySelectorAll('input[type=checkbox]').forEach(cb => cb.checked = false);
  updateStats();
  applyClip();
}}

function selBest(n) {{
  const sorted = [...seriesMeta].filter(m => m.avg != null).sort((a,b) => a.avg - b.avg);
  applySelection(sorted.slice(0, n));
}}

function selWorst(n) {{
  const sorted = [...seriesMeta].filter(m => m.avg != null).sort((a,b) => b.avg - a.avg);
  applySelection(sorted.slice(0, n));
}}

function applySelection(selected) {{
  const ids = new Set(selected.map(m => m.id));
  const sel = {{}};
  legendData.forEach(name => {{
    const meta = seriesMeta.find(m => {{
      const ep = DATA[m.regionIdx].endpoints[m.epIdx];
      return (ep.label + ' (' + ep.label_cn + ')') === name;
    }});
    sel[name] = meta ? ids.has(meta.id) : false;
  }});
  chart.setOption({{ legend: {{ selected: sel }} }});
  document.querySelectorAll('input[type=checkbox]').forEach(cb => cb.checked = false);
  seriesMeta.forEach(m => {{
    if (ids.has(m.id)) {{
      const cb = document.querySelectorAll('#list_' + m.regionIdx + ' .ep-row input[type=checkbox]')[m.epIdx];
      if (cb) cb.checked = true;
    }}
  }});
  updateStats();
  applyClip();
}}

function byRegion() {{
  // Toggle between showing all endpoints and showing only region averages
  const btn = document.querySelector('button[onclick="byRegion()"]');
  const showingRegionAvg = btn.classList.contains('on');
  if (showingRegionAvg) {{
    btn.classList.remove('on');
    selAll();
  }} else {{
    btn.classList.add('on');
    const sel = {{}};
    legendData.forEach(n => sel[n] = n.startsWith('◆'));
    chart.setOption({{ legend: {{ selected: sel }} }});
    document.querySelectorAll('input[type=checkbox]').forEach(cb => cb.checked = false);
    document.querySelectorAll('.region-hdr input[type=checkbox]').forEach(cb => cb.checked = true);
    updateStats();
    applyClip();
  }}
}}

function resetChart() {{
  chart.dispatchAction({{ type: 'restore' }});
  selAll();
}}

// ---- Clip extremes ----
let clipEnabled = false;
function toggleClip() {{
  clipEnabled = !clipEnabled;
  const btn = document.getElementById('btn-clip');
  btn.classList.toggle('on', clipEnabled);
  applyClip();
}}

function applyClip() {{
  if (!clipEnabled) {{
    chart.setOption({{ yAxis: {{ max: null }} }});
    return;
  }}
  // Collect all visible latency values
  const sel = chart.getOption().legend[0].selected;
  const allVals = [];
  series.forEach((s, i) => {{
    const name = legendData[i];
    if (sel[name] !== false && s.data && s.data.length > 0) {{
      s.data.forEach(v => allVals.push(v));
    }}
  }});
  if (allVals.length === 0) return;
  allVals.sort((a,b) => a - b);
  const p95 = allVals[Math.floor(allVals.length * 0.95)];
  // Set Y max to p95, but at least 100ms
  chart.setOption({{ yAxis: {{ max: Math.max(p95, 100) }} }});
}}

// Sync legend clicks with sidebar
chart.on('legendselectchanged', function() {{
  setTimeout(() => {{
    const sel = chart.getOption().legend[0].selected;
    document.querySelectorAll('.ep-row input[type=checkbox]').forEach(cb => {{
      const row = cb.closest('.ep-row');
      const spans = row.querySelectorAll('span');
      if (spans.length >= 2) {{
        const name = spans[1].textContent;
        const cn = row.querySelector('.cn');
        const fullName = name + ' (' + (cn ? cn.textContent : '') + ')';
        cb.checked = sel[fullName] !== false;
      }}
    }});
    updateStats();
  applyClip();
  }}, 50);
}});

// Resize
window.addEventListener('resize', () => chart.resize());

// Init stats
updateStats();
  applyClip();
</script>
</body>
</html>'''

    output_path.write_text(html, encoding="utf-8")
    return output_path

# ===========================================================================
#  MAIN
# ===========================================================================

def main():
    results = []
    total = len(ENDPOINTS)
    print(f"Oracle Cloud OCI Latency Probe — {total} endpoints, {PROBE_COUNT} probes each\n")

    for idx, (region, name_cn, name_en, host) in enumerate(ENDPOINTS, 1):
        print(f"[{idx:2d}/{total}] Probing {name_en} ({region}) -> {host}:{PORT}  ...", end=" ", flush=True)
        latencies = probe_host(host, count=PROBE_COUNT)
        received = len(latencies)
        lost = PROBE_COUNT - received

        if received == 0:
            print("ALL FAILED")
            results.append(dict(region=region, name_cn=name_cn, name_en=name_en, host=host,
                sent=PROBE_COUNT, received=0, lost=lost, loss_pct=100.0,
                min_ms=None, max_ms=None, avg_ms=None, trimmed_avg_ms=None, latencies=[]))
            continue

        mn, mx = round(min(latencies),2), round(max(latencies),2)
        avg = round(statistics.mean(latencies),2)
        ta = trimmed_mean(latencies)
        print(f"min={mn}ms  max={mx}ms  avg={avg}ms  trimmed_avg={ta}ms")
        results.append(dict(region=region, name_cn=name_cn, name_en=name_en, host=host,
            sent=PROBE_COUNT, received=received, lost=lost,
            loss_pct=round(lost/PROBE_COUNT*100,1),
            min_ms=mn, max_ms=mx, avg_ms=avg, trimmed_avg_ms=ta, latencies=latencies))
        time.sleep(1.0)  # gentle gap between endpoints to avoid router flood protection

    xlsx_path = Path.cwd() / "oci_latency_report.xlsx"
    build_excel(results, xlsx_path)
    print(f"\nExcel report saved to: {xlsx_path}")

    html_path = Path.cwd() / "oci_latency_chart.html"
    build_chart_html(results, html_path)
    print(f"Interactive chart saved to: {html_path}")
    print("\nDone!")

if __name__ == "__main__":
    main()